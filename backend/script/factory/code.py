# -*- coding: UTF-8 -*-
"""
@Project : fbapy
@File    : code.py
@Author  : guhua@jiqid.com
@Date    : 2026/02/03 10:35
"""
import re

import openpyxl
from pathlib import Path
from typing import List, Union, Tuple

from backend.common.log import log
from backend.common.security.auth import identity_verifier


class Code:
    @staticmethod
    def write_to_excel(data: List[Tuple[str, str, str]], output_path: Union[str, Path]) -> bool:
        """将数据写入Excel文件 """
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active

            # 写入表头
            headers = ["WifiMAC", "did", "Key"]
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)

            # 写入数据
            for row_idx, (mac, did, key) in enumerate(data, start=2):
                # 格式化MAC地址（添加冒号分隔）
                formatted_mac = ':'.join(re.findall('..', mac))
                sheet.cell(row=row_idx, column=1, value=formatted_mac)
                sheet.cell(row=row_idx, column=2, value=did)
                sheet.cell(row=row_idx, column=3, value=key)

            wb.save(output_path)

            log.info(f"成功 写入 到{output_path} {len(data)} 条数据")
            return True
        except Exception as e:
            log.error(f"写入Excel失败: {str(e)}")
            return False

    @staticmethod
    def read_from_excel(input_path: Union[str, Path]) -> List[Tuple[str, str, str]]:
        """从Excel文件读取数据 """
        try:
            wb = openpyxl.load_workbook(input_path, data_only=True)
            sheet = wb.active

            data = []

            # 从第二行开始读取（跳过表头）
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # 跳过空行
                    continue

                mac, did, key = row[:3]  # 只取前三列

                # 如果MAC地址中有冒号分隔符，去掉它们（保持原始格式）
                if mac and isinstance(mac, str):
                    mac = mac.replace(':', '').replace('-', '').upper()

                # 确保所有值都是字符串
                mac = str(mac) if mac else ""
                did = str(did) if did else ""
                key = str(key) if key else ""

                data.append((mac, did, key))

            log.info(f"从 {input_path} 成功读取 {len(data)} 条数据")
            return data

        except FileNotFoundError:
            log.error(f"文件不存在: {input_path}")
            return []
        except Exception as e:
            log.error(f"读取Excel文件失败: {str(e)}")
            return []

    @classmethod
    def process(cls, input_file: Union[str, Path], output_file: Union[str, Path]) -> bool:
        original_data = cls.read_from_excel(input_file)

        processed_data = []
        for idx, (mac, old_did, old_key) in enumerate(original_data, start=1):
            credentials = identity_verifier.derive_credentials(mac)
            did = credentials["did"]
            key = credentials["key"]

            processed_data.append((mac, did, key))

            if idx % 100 == 0:
                log.info(f"已处理 {idx} 条记录...")

        return cls.write_to_excel(processed_data, output_file)


if __name__ == '__main__':
    input_file = "xlsx/小米三元组_20260203_103140_K11_200.xlsx"
    output_file = "xlsx/小米三元组_20260203_103140_K11_200_output.xlsx"
    Code.process(input_file, output_file)
